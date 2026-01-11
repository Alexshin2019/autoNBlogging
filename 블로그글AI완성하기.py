# -*- coding: utf-8 -*-
import os
import sys
from openpyxl import load_workbook

# Windows 콘솔 인코딩 설정
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8')

from google import genai


class BlogContentGenerator:
    """블로그 본문을 자동으로 생성하는 클래스"""
    
    def __init__(self, api_key):
        """
        Gemini API 클라이언트 초기화
        
        Args:
            api_key: Gemini API 키
        """
        try:
            self.client = genai.Client(api_key=api_key)
            self.model_name = "gemini-2.0-flash-exp"
            print(f"[OK] Gemini API 클라이언트 초기화 완료 (모델: {self.model_name})")
        except Exception as error:
            raise Exception(f"Gemini API 초기화 실패: {str(error)}")
    
    def generate_blog_content(self, title):
        """
        블로그 제목을 입력받아 서론-본론-결론 구조의 본문 생성
        
        Args:
            title: 블로그 제목
            
        Returns:
            생성된 블로그 본문 텍스트
        """
        prompt = f"""
다음 제목으로 블로그 본문을 작성해주세요:

제목: {title}

요구사항:
1. 서론-본론-결론의 3단 구조로 작성할 것
2. 서론: 주제를 소개하고 독자의 관심을 끌 것
3. 본론: 핵심 내용을 구체적이고 상세하게 설명할 것 (2-3개의 소주제 포함)
4. 결론: 내용을 요약하고 행동을 유도할 것
5. 읽기 쉽고 자연스러운 한국어로 작성할 것
6. 실용적인 정보와 구체적인 예시를 포함할 것
7. 약 800-1200자 분량으로 작성할 것
8. 마크다운 문법을 사용하지 말고 순수 텍스트로만 작성할 것

블로그 본문을 작성해주세요:
"""
        
        try:
            # 생성 설정
            generation_config = {
                "temperature": 1.0,
                "max_output_tokens": 8192,
            }
            
            # API 호출
            response = self.client.models.generate_content(
                model=self.model_name,
                contents=prompt,
                config=generation_config
            )
            
            # 응답 텍스트 추출
            result_text = response.text
            return result_text
            
        except Exception as error:
            raise Exception(f"블로그 본문 생성 실패: {str(error)}")


def process_blog_titles(excel_file_path, api_key):
    """
    엑셀 파일의 제목을 읽어 블로그 본문을 생성하고 저장
    
    Args:
        excel_file_path: 엑셀 파일 경로
        api_key: Gemini API 키
    """
    try:
        # Gemini API 초기화
        generator = BlogContentGenerator(api_key)
        
        # 엑셀 파일 열기
        print(f"\n[파일 열기] {excel_file_path}")
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
        
        # 데이터가 있는 마지막 행 찾기
        max_row = sheet.max_row
        print(f"[정보] 총 {max_row - 1}개의 제목 발견 (헤더 제외)")
        
        # 2행부터 마지막 행까지 순회 (1행은 헤더)
        for row_index in range(2, max_row + 1):
            # A열에서 제목 읽기
            title_cell = sheet.cell(row=row_index, column=1)
            title = title_cell.value
            
            # 제목이 비어있으면 건너뛰기
            if not title or str(title).strip() == "":
                print(f"[건너뛰기] {row_index}행: 제목이 비어있음")
                continue
            
            print(f"\n{'='*60}")
            print(f"[진행중] 현재 {row_index}행: {title}")
            print(f"{'='*60}")
            
            try:
                # 블로그 본문 생성
                blog_content = generator.generate_blog_content(title)
                
                # B열에 생성된 본문 저장
                content_cell = sheet.cell(row=row_index, column=2)
                content_cell.value = blog_content
                
                print(f"[완료] {row_index}행 본문 생성 완료 (길이: {len(blog_content)}자)")
                
            except Exception as error:
                # 예외 발생 시 에러 메시지 출력하고 다음 행으로 진행
                print(f"[ERROR] {row_index}행 처리 중 오류 발생: {str(error)}")
                print(f"[계속] 다음 행으로 진행합니다...")
                continue
        
        # 수정된 데이터를 원본 파일에 덮어쓰기
        print(f"\n[저장중] 파일 저장: {excel_file_path}")
        try:
            workbook.save(excel_file_path)
            print(f"[완료] 모든 작업이 완료되었습니다!")
        except PermissionError:
            # 파일이 열려있는 경우 백업 파일로 저장
            backup_file_path = excel_file_path.replace('.xlsx', '_완성본.xlsx')
            workbook.save(backup_file_path)
            print(f"[경고] 원본 파일이 열려있어 저장할 수 없습니다.")
            print(f"[완료] 대신 새 파일로 저장했습니다: {backup_file_path}")
            print(f"[안내] 원본 파일을 닫은 후 다시 실행하거나, 완성본 파일을 사용하세요.")
        
    except FileNotFoundError:
        print(f"[ERROR] 파일을 찾을 수 없습니다: {excel_file_path}")
    except Exception as error:
        print(f"[ERROR] 파일 처리 중 오류 발생: {str(error)}")


def main():
    """메인 실행 함수"""
    # API 키 설정 (실제 사용 시 환경변수 사용 권장)
    api_key = "AIzaSyD1Iih0tQ528MxCBhV8yhmi7wHBc7nZnac"
    
    # 엑셀 파일 경로
    excel_file_path = "posting.xlsx"
    
    # 현재 디렉토리에 파일이 없다면 절대 경로로 시도
    if not os.path.exists(excel_file_path):
        excel_file_path = os.path.join(os.path.dirname(__file__), "posting.xlsx")
    
    print("="*60)
    print("블로그 글 AI 자동 완성 프로그램")
    print("="*60)
    
    # 블로그 제목 처리
    process_blog_titles(excel_file_path, api_key)


if __name__ == "__main__":
    main()
