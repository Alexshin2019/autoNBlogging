# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import pyperclip
import time
from datetime import datetime
import openpyxl

# 네이버 계정 정보
NAVER_ID = "shinung"
NAVER_PW = "wE0905**"

# 엑셀 파일 경로
EXCEL_FILE = "posting_완성본.xlsx"

# 로그 파일 설정
log_file = f"엑셀자동화로그_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

def log_print(message):
    """화면에 출력하고 로그 파일에도 저장"""
    print(message)
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(message + "\n")

def setup_driver():
    """크롬 드라이버 초기화"""
    options = webdriver.ChromeOptions()
    # 자동화 탐지 우회 옵션
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    # 비밀번호 저장 팝업 제거
    prefs = {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    }
    options.add_experimental_option("prefs", prefs)
    
    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    return driver

def input_with_clipboard(element, text):
    """클립보드를 이용해 텍스트 입력"""
    element.click()
    time.sleep(0.3)
    
    # 기존 내용 전체 선택 후 삭제
    element.send_keys(Keys.CONTROL + 'a')
    time.sleep(0.1)
    element.send_keys(Keys.DELETE)
    time.sleep(0.1)
    
    # 클립보드에 복사 후 붙여넣기
    pyperclip.copy(text)
    time.sleep(0.1)
    element.send_keys(Keys.CONTROL + 'v')
    time.sleep(0.3)

def read_excel_data():
    """엑셀 파일에서 데이터 읽기 (2행부터 마지막 행까지)"""
    try:
        log_print(f"\n엑셀 파일 읽기 시작: {EXCEL_FILE}")
        
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        
        # 데이터 리스트 생성
        posts = []
        
        # 2행부터 마지막 행까지 읽기
        for row_num in range(2, sheet.max_row + 1):
            title_cell = sheet.cell(row=row_num, column=1)  # A열
            content_cell = sheet.cell(row=row_num, column=2)  # B열
            
            title = title_cell.value
            content = content_cell.value
            
            # 제목과 본문이 모두 있는 경우만 추가
            if title and content:
                title = str(title).strip()
                content = str(content).strip()
                posts.append({
                    "row": row_num,
                    "title": title,
                    "content": content
                })
                log_print(f"  [{row_num}행] 제목: {title[:30]}...")
            else:
                log_print(f"  [{row_num}행] 건너뜀 (빈 데이터)")
        
        workbook.close()
        
        log_print(f"\n총 {len(posts)}개의 글을 읽었습니다.")
        return posts
        
    except FileNotFoundError:
        log_print(f"[ERROR] 엑셀 파일을 찾을 수 없습니다: {EXCEL_FILE}")
        log_print("현재 디렉토리에 posting.xlsx 파일이 있는지 확인하세요.")
        return []
    except Exception as e:
        log_print(f"[ERROR] 엑셀 파일 읽기 오류: {str(e)}")
        return []

def naver_login(driver):
    """네이버 로그인 수행"""
    try:
        log_print("\n네이버 로그인 페이지 접속 중...")
        driver.get("https://nid.naver.com/nidlogin.login")
        time.sleep(2)
        
        # 아이디 입력
        log_print("아이디 입력 중...")
        id_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "id"))
        )
        input_with_clipboard(id_input, NAVER_ID)
        
        # 비밀번호 입력
        log_print("비밀번호 입력 중...")
        pw_input = driver.find_element(By.ID, "pw")
        input_with_clipboard(pw_input, NAVER_PW)
        
        # 로그인 버튼 클릭
        log_print("로그인 버튼 클릭...")
        login_button = driver.find_element(By.ID, "log.login")
        login_button.click()
        
        time.sleep(3)
        log_print("로그인 완료!")
        
        return True
        
    except Exception as e:
        log_print(f"[ERROR] 로그인 오류: {str(e)}")
        return False

def close_popups(driver):
    """팝업 닫기"""
    try:
        close_button_selectors = [
            ".se-popup-button-cancel",
            ".se-help-panel-close-button",
            "[class*='popup'] button[class*='close']",
            "button[aria-label*='닫기']",
        ]
        
        for selector in close_button_selectors:
            try:
                buttons = driver.find_elements(By.CSS_SELECTOR, selector)
                for btn in buttons:
                    if btn.is_displayed():
                        driver.execute_script("arguments[0].click();", btn)
                        time.sleep(0.3)
            except:
                continue
        
        # ESC 키로도 시도
        try:
            actions = ActionChains(driver)
            actions.send_keys(Keys.ESCAPE)
            actions.perform()
            time.sleep(0.3)
        except:
            pass
            
    except Exception as e:
        pass

def write_single_post(driver, title, content):
    """블로그 글 하나 작성"""
    try:
        # 블로그 글쓰기 페이지로 이동
        log_print("\n블로그 글쓰기 페이지로 이동...")
        driver.get("https://blog.naver.com/GoBlogWrite.naver")
        time.sleep(5)
        
        # iframe 전환
        log_print("iframe으로 전환 중...")
        iframe = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#mainFrame"))
        )
        driver.switch_to.frame(iframe)
        time.sleep(3)
        
        # 팝업 닫기
        for _ in range(3):
            close_popups(driver)
            time.sleep(0.5)
        
        # 팝업 오버레이 제거
        try:
            driver.execute_script("""
                var popups = document.querySelectorAll('[class*="popup"], [class*="modal"], [class*="overlay"]');
                popups.forEach(function(el) {
                    el.style.display = 'none';
                    el.style.visibility = 'hidden';
                });
            """)
        except:
            pass
        
        time.sleep(1)
        
        # 제목 입력
        log_print(f"제목 입력: {title[:50]}...")
        try:
            title_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".se-section-documentTitle"))
            )
            
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", title_element)
            time.sleep(1)
            
            # JavaScript로 직접 텍스트 입력 및 글씨 크기 설정
            driver.execute_script("""
                var element = arguments[0];
                element.focus();
                element.textContent = arguments[1];
                element.innerText = arguments[1];
                
                // 제목 글씨 크기 크게 설정 (26px)
                element.style.fontSize = '26px';
                element.style.fontWeight = 'bold';
                
                var event = new Event('input', { bubbles: true });
                element.dispatchEvent(event);
            """, title_element, title)
            
            log_print("  제목 입력 완료 (글씨 크기: 26px)")
            time.sleep(1)
                
        except Exception as e:
            log_print(f"  제목 입력 오류: {str(e)}")
        
        # 본문 입력
        log_print(f"본문 입력 중... ({len(content)}자)")
        try:
            content_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".se-section-text"))
            )
            
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", content_element)
            time.sleep(1)
            
            # 본문 클릭 및 포커스
            driver.execute_script("""
                var element = arguments[0];
                element.click();
                element.focus();
            """, content_element)
            time.sleep(1)
            
            # 본문을 문단별로 나누어 입력
            paragraphs = [p.strip() for p in content.split('\n') if p.strip()]
            log_print(f"  총 {len(paragraphs)}개 문단 입력...")
            
            for i, paragraph in enumerate(paragraphs):
                # ActionChains로 타이핑
                actions = ActionChains(driver)
                for char in paragraph:
                    actions.send_keys(char)
                    actions.pause(0.01)
                actions.perform()
                
                # 문단 사이 Enter
                if i < len(paragraphs) - 1:
                    time.sleep(0.2)
                    actions = ActionChains(driver)
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    time.sleep(0.2)
            
            log_print("  본문 입력 완료")
            time.sleep(2)
                
        except Exception as e:
            log_print(f"  본문 입력 오류: {str(e)}")
        
        # 저장 버튼 클릭
        log_print("저장 버튼 검색 중...")
        
        # iframe 내부에서 저장 버튼 찾기
        save_button = None
        try:
            iframe_buttons = driver.find_elements(By.TAG_NAME, "button")
            for btn in iframe_buttons:
                try:
                    btn_text = btn.text.strip()
                    if "임시저장" in btn_text or "저장" in btn_text or "완료" in btn_text:
                        save_button = btn
                        log_print(f"  저장 버튼 발견: '{btn_text}'")
                        break
                except:
                    continue
        except:
            pass
        
        # iframe 내부에서 찾았으면 클릭
        if save_button:
            try:
                driver.execute_script("arguments[0].click();", save_button)
                time.sleep(3)
                log_print("  저장 완료!")
                driver.switch_to.default_content()
                return True
            except Exception as e:
                log_print(f"  저장 버튼 클릭 실패: {str(e)}")
        
        # iframe 밖에서 찾기
        driver.switch_to.default_content()
        time.sleep(3)
        
        # 텍스트로 저장 버튼 찾기
        try:
            all_buttons = driver.find_elements(By.TAG_NAME, "button")
            for btn in all_buttons:
                try:
                    btn_text = btn.text.strip()
                    if btn_text and ("임시저장" in btn_text or "저장" in btn_text):
                        save_button = btn
                        log_print(f"  저장 버튼 발견: '{btn_text}'")
                        break
                except:
                    continue
        except:
            pass
        
        if save_button:
            try:
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", save_button)
                time.sleep(1)
                driver.execute_script("arguments[0].click();", save_button)
                time.sleep(3)
                log_print("  저장 완료!")
                return True
            except Exception as e:
                log_print(f"  저장 버튼 클릭 실패: {str(e)}")
        
        # 저장 버튼을 못 찾은 경우
        log_print("  [주의] 자동 저장 버튼을 찾지 못했습니다.")
        log_print("  수동으로 저장 버튼을 눌러주세요. (5초 대기)")
        time.sleep(5)
        
        return True
        
    except Exception as e:
        log_print(f"[ERROR] 글 작성 중 오류: {str(e)}")
        try:
            driver.switch_to.default_content()
        except:
            pass
        return False

def main():
    """메인 실행 함수"""
    driver = None
    
    try:
        log_print(f"\n{'='*60}")
        log_print(f"엑셀 기반 네이버 블로그 자동 글쓰기 시작")
        log_print(f"로그 파일: {log_file}")
        log_print(f"{'='*60}\n")
        
        # 엑셀 파일에서 데이터 읽기
        posts = read_excel_data()
        
        if not posts:
            log_print("\n[ERROR] 작성할 글이 없습니다.")
            log_print("posting.xlsx 파일을 확인하세요.")
            return
        
        # 드라이버 설정
        driver = setup_driver()
        
        # 네이버 로그인
        if not naver_login(driver):
            log_print("\n[ERROR] 로그인 실패")
            return
        
        # 각 글 작성
        success_count = 0
        fail_count = 0
        
        log_print(f"\n{'='*60}")
        log_print(f"총 {len(posts)}개의 글 작성 시작")
        log_print(f"{'='*60}")
        
        for idx, post in enumerate(posts, 1):
            log_print(f"\n[{idx}/{len(posts)}] {post['row']}행 작성 중...")
            log_print(f"  제목: {post['title'][:50]}")
            
            if write_single_post(driver, post['title'], post['content']):
                success_count += 1
                log_print(f"  ✓ 성공")
            else:
                fail_count += 1
                log_print(f"  ✗ 실패")
            
            # 다음 글 작성 전 대기
            if idx < len(posts):
                log_print("  다음 글 작성까지 3초 대기...")
                time.sleep(3)
        
        # 최종 결과
        log_print(f"\n{'='*60}")
        log_print(f"작업 완료!")
        log_print(f"  성공: {success_count}개")
        log_print(f"  실패: {fail_count}개")
        log_print(f"  전체: {len(posts)}개")
        log_print(f"{'='*60}")
        
        # 사용자 대기
        log_print("\n작업이 끝났습니다.")
        log_print("브라우저를 확인하고 Enter를 눌러 종료하세요...")
        input()
        
    except Exception as e:
        log_print(f"\n[ERROR] 프로그램 실행 중 오류: {str(e)}")
        
    finally:
        # 드라이버 종료
        if driver:
            driver.quit()
            log_print("\n브라우저 종료 완료")
        
        log_print(f"\n{'='*60}")
        log_print(f"모든 로그가 '{log_file}' 파일에 저장되었습니다.")
        log_print(f"{'='*60}")

if __name__ == "__main__":
    main()
